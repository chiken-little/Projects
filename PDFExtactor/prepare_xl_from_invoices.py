from extract_data_from_invoice import *
import openpyxl

wb = openpyxl.Workbook()
wb.title = "Invoice Data"
ws = wb.active
ws['A1'] = "MOBILE"
ws['B1'] = "NAME"
ws['C1'] = "CALLS"
ws['D1'] = "SMS"
ws['E1'] = "TOTAL"

for page in range(npages):
    invoice = invoice_data(page, True)
    if 'mobile' in invoice:
        row = []
        row.append(invoice['mobile'])
        row.append(invoice['name'])
        row.append(str(sum([invoice['Talk Summary'][i] for i in invoice['Talk Summary']])))
        row.append(str(invoice['SMS and MMS Summary']['Total SMS']))
        row.append(str(invoice['Account Summary']['Grand Total']))
        ws.append(row)

wb.save("output.xlsx")
